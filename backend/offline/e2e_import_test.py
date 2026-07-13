"""端到端验证：密码批量导入 + 模板下载 + 导出权限（普通用户亦可）+ 导入后解密。

加密在服务端容器内进行，本机只做 HTTP 客户端，安全。
目标地址通过环境变量 BASE 指定（默认 http://localhost:9012）。
xlsx 直接在测试进程内用 openpyxl 生成，无需 docker exec。
"""
import csv, io, json, os, struct, subprocess, urllib.request, urllib.error
from openpyxl import Workbook
from openpyxl.styles import Font

BASE = os.getenv("BASE", "http://localhost:9012")

# ── SM3（与前端一致，用于 SCRAM 登录）──
_SM3_IV = [0x7380166f, 0x4914b2b9, 0x172442d7, 0xda8a0600,
           0xa96f30bc, 0x163138aa, 0xe38dee4d, 0xb0fb0e4e]
def _rotl(x, n):
    x &= 0xffffffff
    return ((x << n) | (x >> (32 - n))) & 0xffffffff
def _p0(x):
    return x ^ _rotl(x, 9) ^ _rotl(x, 17)
def _p1(x):
    return x ^ _rotl(x, 15) ^ _rotl(x, 23)
def _ff(x, y, z, j):
    return (x ^ y ^ z) if j < 16 else ((x & y) | (x & z) | (y & z))
def _gg(x, y, z, j):
    return (x ^ y ^ z) if j < 16 else ((x & y) | ((~x) & z))
def _tj(j):
    return 0x79cc4519 if j < 16 else 0x7a879d8a
def _cf(V, B):
    W = [0] * 68
    for i in range(16):
        W[i] = struct.unpack(">I", B[i * 4:i * 4 + 4])[0]
    for i in range(16, 68):
        x = W[i - 16] ^ W[i - 9] ^ _rotl(W[i - 3], 15)
        W[i] = _p1(x) ^ _rotl(W[i - 13], 7) ^ W[i - 6]
    W1 = [W[i] ^ W[i + 4] for i in range(64)]
    A, B2, C, D, E, F, G, H = V
    for j in range(64):
        SS1 = _rotl(_rotl(A, 12) + E + _rotl(_tj(j), j % 32), 7) & 0xffffffff
        SS2 = SS1 ^ _rotl(A, 12)
        TT1 = (_ff(A, B2, C, j) + D + SS2 + W1[j]) & 0xffffffff
        TT2 = (_gg(E, F, G, j) + H + SS1 + W[j]) & 0xffffffff
        D = C; C = _rotl(B2, 9); B2 = A; A = TT1
        H = G; G = _rotl(F, 19); F = E; E = _p0(TT2)
    return [V[i] ^ v for i, v in enumerate([A, B2, C, D, E, F, G, H])]
def sm3_hex(data):
    V = _SM3_IV[:]
    bl = len(data) * 8
    msg = bytearray(data); msg.append(0x80)
    while len(msg) % 64 != 56:
        msg.append(0)
    msg += struct.pack(">Q", bl)
    for i in range(0, len(msg), 64):
        V = _cf(V, msg[i:i + 64])
    return b"".join(struct.pack(">I", v) for v in V).hex()

def req(method, path, body=None, token=None, headers=None, raw=False):
    url = BASE + path
    h = {}
    if token:
        h["Authorization"] = "Bearer " + token
    data = None
    if body is not None:
        if isinstance(body, (bytes, bytearray)):
            data = body
        else:
            data = json.dumps(body).encode("utf-8")
            h["Content-Type"] = "application/json"
    if headers:
        h.update(headers)
    r = urllib.request.Request(url, data=data, method=method, headers=h)
    try:
        resp = urllib.request.urlopen(r, timeout=30)
        rb = resp.read()
        return resp.status, (rb if raw else (json.loads(rb) if rb else None))
    except urllib.error.HTTPError as e:
        return e.code, (e.read() if raw else None)

def scram_login(username, password):
    st, chal = req("POST", "/api/auth/login/begin", {"username": username})
    if st == 409:
        st, b = req("POST", "/api/auth/login", {"username": username, "password": password})
        return b["access_token"]
    T = sm3_hex(password.encode("utf-8") + bytes.fromhex(chal["salt"]))
    proof = sm3_hex(bytes.fromhex(T) + bytes.fromhex(chal["nonce"]))
    st, b = req("POST", "/api/auth/login/verify",
                   {"username": username, "nonce": chal["nonce"], "proof": proof})
    return b["access_token"]

def multipart(fields, file_tuple):
    boundary = "----e2eboundary1234"
    body = b""
    for k, v in fields.items():
        body += f"--{boundary}\r\n".encode()
        body += f'Content-Disposition: form-data; name="{k}"\r\n\r\n'.encode()
        body += str(v).encode() + b"\r\n"
    name, fname, fdata, ctype = file_tuple
    body += f"--{boundary}\r\n".encode()
    body += f'Content-Disposition: form-data; name="{name}"; filename="{fname}"\r\n'.encode()
    body += f"Content-Type: {ctype}\r\n\r\n".encode()
    body += fdata + b"\r\n"
    body += f"--{boundary}--\r\n".encode()
    return body, f"multipart/form-data; boundary={boundary}"

print("=== 1. 管理员登录 ===")
admin = scram_login("admin", "TestPass!2026")
print("admin token len:", len(admin))

print("\n=== 2. 下载导入模板（xlsx / csv） ===")
st, xlsx = req("GET", "/api/passwords/template?fmt=xlsx", token=admin, raw=True)
print("xlsx template:", st, "bytes", len(xlsx))
open("/tmp/pwd_tpl.xlsx", "wb").write(xlsx)
st, csvb = req("GET", "/api/passwords/template?fmt=csv", token=admin, raw=True)
print("csv template:", st, "bytes", len(csvb), "(期望 400，已废弃 csv)")
assert st == 400, "csv 模板已废弃，应返回 400"

print("\n=== 3. 取一个存在的分组名 ===")
st, groups = req("GET", "/api/admin/groups", token=admin)
gname = groups[0]["name"]
gid = groups[0]["id"]
print("使用分组:", gname, "(id=%s)" % gid)

print("\n=== 4. 非管理员导出应可达（Req1：普通用户也能导出，不再 403） ===")
# 创建非管理员
st, cu = req("POST", "/api/admin/users",
             {"username": "imp_tester", "password": "ImpTester!1", "is_admin": False, "group_ids": [gid]},
             token=admin)
print("创建测试用户:", st)
uid = cu["id"] if st == 200 else None
tester = scram_login("imp_tester", "ImpTester!1")
# 空 ids 会触发 400（参数校验），但绝不应是 403（管理员门禁已移除）
st, _ = req("POST", "/api/passwords/export", {"ids": [], "format": "json", "plaintext": True}, token=tester)
print("非管理员导出(空) ->", st, "(期望 200 或 400，但不再是 403)")
assert st != 403, "非管理员不应被禁止导出（Req1 已放开导出权限）"

print("\n=== 5. 管理员导出可达（空 ids -> 400 参数校验） ===")
st, _ = req("POST", "/api/passwords/export", {"ids": [], "format": "json", "plaintext": True}, token=admin)
print("管理员导出(空) ->", st, "(期望 400，参数校验而非权限校验)")

print("\n=== 6. xlsx 批量导入（表头：标题/账号/密码明文/备注/所属分组） ===")
wb = Workbook(); ws = wb.active; ws.title = "t"
ws.append(["标题", "账号", "密码明文", "备注", "所属分组"])
ws.append(["网站A", "imp_a", "SecretA!1", "导入测试", gname])
ws.append(["网站B", "imp_b", "SecretB!2", "导入测试", gname])
wb.save("/tmp/imp6.xlsx")
with open("/tmp/imp6.xlsx", "rb") as f:
    xlsx6 = f.read()
body, ctype = multipart(
    {"algorithm": "symmetric", "entry_password": "ImpPass!2026", "group_id": str(gid)},
    ("file", "import_test.xlsx", xlsx6, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
)
st, resp = req("POST", "/api/passwords/import", body=body, token=admin,
               headers={"Content-Type": ctype})
print("导入 ->", st)
print(json.dumps(resp, ensure_ascii=False) if resp else "")
assert st == 200 and resp["created"] == 2, "应有 2 条创建成功"

print("\n=== 7. 导入后解密验证（确认内层用导入的 entry_password） ===")
st, entries = req("GET", "/api/passwords", token=admin)
imp_a = next((e for e in entries if e["username"] == "imp_a"), None)
assert imp_a, "imp_a 未出现在列表"
st, full = req("POST", f"/api/passwords/{imp_a['id']}/unlock", {"entry_password": "ImpPass!2026"}, token=admin)
print("解锁 imp_a ->", st, "| secret:", full.get("secret"))
assert st == 200 and full["secret"] == "SecretA!1"

print("\n=== 8. xlsx 导入回环（生成 xlsx 再导入） ===")
wb = Workbook(); ws = wb.active; ws.title = "t"
ws.cell(row=1, column=1, value="使用说明").font = Font(bold=True)
ws.cell(row=2, column=1, value="说明行2")
hdr = 5
for c, h in enumerate(["标题", "账号", "密码明文", "备注", "所属分组"], start=1):
    ws.cell(row=hdr, column=c, value=h)
data = [["网站C", "imp_c", "SecretC!3", "xlsx测试", gname],
        ["网站D", "imp_d", "SecretD!4", "xlsx测试", gname]]
for r, row in enumerate(data, start=hdr + 1):
    for c, v in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=v)
wb.save("/tmp/imp.xlsx")
with open("/tmp/imp.xlsx", "rb") as f:
    xlsx_data = f.read()
body, ctype = multipart(
    {"algorithm": "symmetric", "entry_password": "ImpPass!2026", "group_id": str(gid)},
    ("file", "imp.xlsx", xlsx_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
)
st, resp = req("POST", "/api/passwords/import", body=body, token=admin,
               headers={"Content-Type": ctype})
print("xlsx 导入 ->", st)
print(json.dumps(resp, ensure_ascii=False) if resp else "")
assert st == 200 and resp["created"] == 2, "xlsx 应有 2 条创建成功"

print("\n=== 9. 清理测试数据 ===")
for uname in ["imp_a", "imp_b", "imp_c", "imp_d"]:
    st, entries = req("GET", "/api/passwords", token=admin)
    for e in entries:
        if e["username"] == uname:
            req("DELETE", f"/api/passwords/{e['id']}", token=admin)
            print("  删除", uname)
if uid:
    req("DELETE", f"/api/admin/users/{uid}", token=admin)
    print("  删除测试用户 imp_tester")

print("\nALL E2E PASS ✅")
