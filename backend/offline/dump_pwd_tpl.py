import ast, io, sys

src = open("/Users/liuyupengliu/Downloads/projects/pwd-web-new/backend/app/routers/passwords.py", encoding="utf-8").read()
tree = ast.parse(src)

# 抽取 PWD_IMPORT_HEADERS 常量 与 _xlsx_bytes_passwords 函数定义
ns = {"io": io}
for node in tree.body:
    if isinstance(node, ast.Assign):
        for t in node.targets:
            if isinstance(t, ast.Name) and t.id == "PWD_IMPORT_HEADERS":
                ns["PWD_IMPORT_HEADERS"] = ast.literal_eval(node.value)
    if isinstance(node, ast.FunctionDef) and node.name == "_xlsx_bytes_passwords":
        code = compile(ast.Module(body=[node], type_ignores=[]), "<fn>", "exec")
        exec(code, ns)

data = ns["_xlsx_bytes_passwords"]()
open("/tmp/pwd_tpl.xlsx", "wb").write(data)

from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(data))
ws = wb.active
print("SHEET:", ws.title, "dims:", ws.dimensions)
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    cells = [c for c in row if c is not None]
    if cells:
        print(i, cells)
