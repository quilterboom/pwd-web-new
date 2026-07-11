import requests, io
from openpyxl import load_workbook

BASE = "http://localhost:9019"

def auth(u, p):
    r = requests.post(f"{BASE}/api/auth/login", json={"username": u, "password": p}, timeout=30)
    assert r.status_code == 200, (u, r.status_code, r.text)
    return r.json()["access_token"]

def dump_xlsx(b):
    wb = load_workbook(io.BytesIO(b))
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [c for c in row if c is not None]
        if cells:
            rows.append((i, cells))
    return rows

tok = auth("admin", "TestPass!2026")

# 密码模板
r1 = requests.get(f"{BASE}/api/passwords/template", params={"fmt": "xlsx"}, headers={"Authorization": f"Bearer {tok}"}, timeout=60)
print("== 密码模板 ==")
print("status", r1.status_code, "bytes", len(r1.content))
for ln, cells in dump_xlsx(r1.content)[:15]:
    print(f"  L{ln}: {cells}")

# 用户模板
r2 = requests.get(f"{BASE}/api/admin/users/template", headers={"Authorization": f"Bearer {tok}"}, timeout=60)
print("\n== 用户模板 ==")
print("status", r2.status_code, "bytes", len(r2.content))
for ln, cells in dump_xlsx(r2.content)[:15]:
    print(f"  L{ln}: {cells}")

# 断言：两者首行说明不同、表头不同
pw_rows = dump_xlsx(r1.content)
user_rows = dump_xlsx(r2.content)
pw_header = next((c for _, c in pw_rows if "密码明文" in c), None)
user_header = next((c for _, c in user_rows if "是否管理员" in c), None)
assert pw_header, "密码模板缺少正确表头"
assert user_header, "用户模板缺少正确表头"
assert pw_header != user_header, "两个模板表头不应相同"
print("\n[PASS] 密码模板与用户模板内容已正确区分")
