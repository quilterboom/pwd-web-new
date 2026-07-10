import sys, json, io, hashlib, urllib.request, urllib.error

BASE = sys.argv[1] if len(sys.argv) > 1 else "http://localhost:9021"
ok = 0; fail = 0

def req(method, path, body=None, token=None, binary=False):
    data = json.dumps(body).encode() if body is not None else None
    h = {"Content-Type": "application/json"}
    if token: h["Authorization"] = "Bearer " + token
    r = urllib.request.Request(BASE + path, data=data, headers=h, method=method)
    try:
        resp = urllib.request.urlopen(r, timeout=15)
        raw = resp.read()
        return resp.status, raw
    except urllib.error.HTTPError as e:
        return e.code, e.read()

def sm3_hex(b):
    h = hashlib.new("sm3"); h.update(b); return h.hexdigest()

def login(username, password):
    st, b = req("POST", "/api/auth/login/begin", {"username": username})
    if st == 409:
        st, b = req("POST", "/api/auth/login", {"username": username, "password": password})
        return json.loads(b)["access_token"]
    ch = json.loads(b)
    T = sm3_hex(password.encode() + bytes.fromhex(ch["salt"]))
    proof = sm3_hex(bytes.fromhex(T) + bytes.fromhex(ch["nonce"]))
    st, b = req("POST", "/api/auth/login/verify",
                 {"username": username, "nonce": ch["nonce"], "proof": proof})
    return json.loads(b)["access_token"]

def check(name, cond):
    global ok, fail
    if cond: ok += 1; print(f"  [PASS] {name}")
    else: fail += 1; print(f"  [FAIL] {name}")

def count_xlsx_rows(content):
    """用 openpyxl 读取 xlsx 数据行数（不含表头）。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    wb.close()
    return max(0, len(rows) - 1)  # 减去表头

print("=== 登录 admin ===")
tok = login("admin", "TestPass!2026")
check("admin 登录", bool(tok))

print("=== 准备：建两个 symmetric 条目（不同解密密码）===")
st, b = req("GET", "/api/admin/groups", token=tok)
g1 = json.loads(b)[0]["id"]
def add_pw(username, ep):
    st, b = req("POST", "/api/passwords",
        {"title": username, "username": username, "secret": "s3", "algorithm": "symmetric",
         "group_id": g1, "entry_password": ep, "comment": "新增密码"}, token=tok)
    return json.loads(b)["id"] if st == 200 else None
id1 = add_pw("exp_a", "pwAAAA1111")
id2 = add_pw("exp_b", "pwBBBB2222")
check("建 exp_a", id1 is not None)
check("建 exp_b", id2 is not None)

print("=== 问题1：解密失败则禁止导出并提示 ===")
# 两个条目用同一错误密码 → 两条都解密失败 → 应 400 拒绝
st, b = req("POST", "/api/passwords/export",
    {"ids": [id1, id2], "passwords": {str(id1): "WRONG", str(id2): "WRONG"},
     "format": "xlsx", "plaintext": True}, token=tok)
check("全错密码导出 = 400", st == 400)
check("400 文案含『无法导出明文』", "无法导出明文" in b.decode("utf-8", "replace"))
check("400 文案含失败账号 exp_a", "exp_a" in b.decode("utf-8", "replace"))

# 其中一个对、一个错 → 只要有失败就整体拒绝
st, b = req("POST", "/api/passwords/export",
    {"ids": [id1, id2], "passwords": {str(id1): "pwAAAA1111", str(id2): "WRONG"},
     "format": "xlsx", "plaintext": True}, token=tok)
check("部分错 → 400 拒绝", st == 400)

# 全部正确 → 200 正常导出 xlsx
st, b = req("POST", "/api/passwords/export",
    {"ids": [id1, id2], "passwords": {str(id1): "pwAAAA1111", str(id2): "pwBBBB2222"},
     "format": "xlsx", "plaintext": True}, token=tok)
check("全对 → 200 导出", st == 200)
if st == 200:
    rows = count_xlsx_rows(b)
    check("导出 xlsx 含 2 条明文（数据行）", rows == 2)
    check("响应为 Excel 二进制（以 PK 头开头）", b[:2] == b"PK")

print("=== 仅支持 xlsx：非 xlsx 格式应被拒绝 ===")
st, b = req("POST", "/api/passwords/export",
    {"ids": [id1, id2], "passwords": {str(id1): "pwAAAA1111", str(id2): "pwBBBB2222"},
     "format": "csv", "plaintext": True}, token=tok)
check("csv 格式 → 400 拒绝", st == 400)
check("400 文案提示仅支持 Excel", "Excel" in b.decode("utf-8", "replace"))

print(f"\n导出解密失败拦截 + xlsx 校验：{ok} 通过 / {fail} 失败")
sys.exit(1 if fail else 0)
