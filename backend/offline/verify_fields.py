"""字段改造端到端校验：密码文件名称/系统/用户名、模板表头、导出表头、system 过滤、非管理员导出。
运行于容器内：docker exec <c> python /verify_fields.py （BASE 为容器内的 9010）
"""
import io, json, struct, sys, urllib.parse, urllib.request, urllib.error
sys.path.insert(0, "/app")
from openpyxl import load_workbook

BASE = "http://localhost:9010"
ADMIN = "admin"
ADMIN_PW = "TestPass!2026"

_SM3_IV = [0x7380166f, 0x4914b2b9, 0x172442d7, 0xda8a0600,
           0xa96f30bc, 0x163138aa, 0xe38dee4d, 0xb0fb0e4e]
def _rotl(x, n):
    x &= 0xffffffff
    return ((x << n) | (x >> (32 - n))) & 0xffffffff
def _p0(x): return x ^ _rotl(x, 9) ^ _rotl(x, 17)
def _p1(x): return x ^ _rotl(x, 15) ^ _rotl(x, 23)
def _ff(x, y, z, j): return (x ^ y ^ z) if j < 16 else ((x & y) | (x & z) | (y & z))
def _gg(x, y, z, j): return (x ^ y ^ z) if j < 16 else ((x & y) | ((~x) & z))
def _tj(j): return 0x79cc4519 if j < 16 else 0x7a879d8a
def _cf(V, B):
    W = [0] * 68
    for i in range(16):
        W[i] = struct.unpack(">I", B[i * 4:i * 4 + 4])[0]
    for i in range(16, 68):
        x = W[i - 16] ^ W[i - 9] ^ _rotl(W[i - 3], 15)
        W[i] = _p1(x) ^ _rotl(W[i - 13], 7) ^ W[i - 6]
    W1 = [W[i] ^ W[i + 4] for i in range(64)]
    A, B2, C, D, E, F, G, H = V
    for j in range(64):
        SS1 = _rotl(_rotl(A, 12) + E + _rotl(_tj(j), j % 32), 7) & 0xffffffff
        SS2 = SS1 ^ _rotl(A, 12)
        TT1 = (_ff(A, B2, C, j) + D + SS2 + W1[j]) & 0xffffffff
        TT2 = (_gg(E, F, G, j) + H + SS1 + W[j]) & 0xffffffff
        D = C; C = _rotl(B2, 9); B2 = A; A = TT1
        H = G; G = _rotl(F, 19); F = E; E = _p0(TT2)
    return [V[i] ^ v for i, v in enumerate([A, B2, C, D, E, F, G, H])]
def sm3_hex(data):
    V = _SM3_IV[:]
    bl = len(data) * 8
    msg = bytearray(data); msg.append(0x80)
    while len(msg) % 64 != 56: msg.append(0)
    msg += struct.pack(">Q", bl)
    for i in range(0, len(msg), 64):
        V = _cf(V, msg[i:i + 64])
    return b"".join(struct.pack(">I", v) for v in V).hex()

def req(method, path, body=None, token=None, raw=False):
    h = {"Content-Type": "application/json"}
    if token: h["Authorization"] = "Bearer " + token
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(BASE + path, data=data, method=method, headers=h)
    try:
        resp = urllib.request.urlopen(r)
        b = resp.read()
        return resp.status, (b if raw else json.loads(b or b"{}"))
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()

def scram_login(username, password):
    st, chal = req("POST", "/api/auth/login/begin", {"username": username})
    if st == 200 and "salt" in chal:
        T = sm3_hex(password.encode("utf-8") + bytes.fromhex(chal["salt"]))
        proof = sm3_hex(bytes.fromhex(T) + bytes.fromhex(chal["nonce"]))
        st, b = req("POST", "/api/auth/login/verify", {"username": username, "nonce": chal["nonce"], "proof": proof})
        if st == 200: return b.get("access_token")
    st, b = req("POST", "/api/auth/login", {"username": username, "password": password})
    return b.get("access_token") if st == 200 else None

def items_of(rows):
    return rows if isinstance(rows, list) else rows.get("items", [])

def user_id_by_name(token, name):
    st, users = req("GET", "/api/admin/users", token=token)
    if st != 200: return None
    u = next((x for x in users if x["username"] == name), None)
    return u["id"] if u else None

print("=== 登录 admin ===")
admin = scram_login(ADMIN, ADMIN_PW)
assert admin, "admin 登录失败"
print("admin token len:", len(admin))

print("\n=== 取默认分组 ===")
st, groups = req("GET", "/api/admin/groups", token=admin)
gid = groups[0]["id"]
print("分组 id =", gid)

print("\n=== 新增带 密码文件名称/系统/用户名 的条目 ===")
st, b = req("POST", "/api/passwords", {
    "title": "邮箱登录密码", "system": "邮箱系统", "username": "verify_alice",
    "secret": "AliceVerify!1", "notes": "验证备注", "group_id": gid,
    "algorithm": "symmetric", "entry_password": "EntryPw!2026",
}, token=admin)
print("新增 ->", st)
assert st == 200, b
pid = b["id"]

print("\n=== 列表按 system 过滤 ===")
st, rows = req("GET", "/api/passwords?system=" + urllib.parse.quote("邮箱系统"), token=admin)
print("system=邮箱系统 ->", st)
found = [e for e in items_of(rows) if e["id"] == pid]
assert found, "system 过滤未命中"
e = found[0]
print("条目字段:", {k: e.get(k) for k in ("title", "system", "username")})
assert e["title"] == "邮箱登录密码" and e["system"] == "邮箱系统" and e["username"] == "verify_alice"

print("\n=== 列表按 group_id 过滤 ===")
st, rows2 = req("GET", f"/api/passwords?group_id={gid}", token=admin)
print("group_id 过滤命中数:", len(items_of(rows2)))

print("\n=== 下载导入模板，检查新表头 ===")
st, tpl = req("GET", "/api/passwords/template?fmt=xlsx", token=admin, raw=True)
print("template ->", st, "bytes", len(tpl))
ws = load_workbook(io.BytesIO(tpl), read_only=True, data_only=True).active
hdr = None
for r in ws.iter_rows(values_only=True):
    row = [c for c in r if c is not None]
    if row and "密码明文" in [str(x).strip() for x in row]:
        hdr = [str(x).strip() for x in row]; break
print("模板表头:", hdr)
for h in ["密码文件名称", "系统", "用户名", "密码明文", "备注"]:
    assert h in hdr, f"模板缺少表头 {h}"
print("模板表头 OK")

print("\n=== 明文导出（admin），检查新表头与字段 ===")
st, resp = req("POST", "/api/passwords/export", {"ids": [pid], "passwords": {str(pid): "EntryPw!2026"}, "format": "xlsx", "plaintext": True}, token=admin, raw=True)
print("export(admin) ->", st, "bytes", len(resp))
ws = load_workbook(io.BytesIO(resp), read_only=True, data_only=True).active
exhdr = None
for r in ws.iter_rows(values_only=True):
    row = [c for c in r if c is not None]
    if row and "密码明文" in [str(x).strip() for x in row]:
        exhdr = [str(x).strip() for x in row]; break
print("导出表头:", exhdr)
for h in ["密码文件名称", "系统", "用户名", "加密方式", "所属分组", "密钥", "密码明文", "备注", "更新时间"]:
    assert h in exhdr, f"导出缺少表头 {h}"
data_row = None
for r in ws.iter_rows(values_only=True):
    if any(str(c).strip() == "邮箱登录密码" for c in r if c is not None):
        data_row = list(r); break
assert data_row, "导出数据行未找到"
print("导出数据:", data_row)
print("导出表头 OK")

print("\n=== Req1：非管理员也能导出 ===")
# 清理可能残留的旧用户
old = user_id_by_name(admin, "verify_normal")
if old:
    req("DELETE", "/api/admin/users/" + str(old), token=admin)
    print("已清理旧 verify_normal")
st, cu = req("POST", "/api/admin/users", {"username": "verify_normal", "password": "VerifyNormal!1", "is_admin": False, "group_ids": [gid]}, token=admin)
print("创建非管理员 ->", st)
ntok = scram_login("verify_normal", "VerifyNormal!1")
assert ntok, "非管理员登录失败"
st, resp2 = req("POST", "/api/passwords/export", {"ids": [pid], "passwords": {str(pid): "EntryPw!2026"}, "format": "xlsx", "plaintext": True}, token=ntok, raw=True)
print("export(非管理员) ->", st)
assert st == 200, f"非管理员导出应成功，实际 {st}"
print("非管理员导出 OK")
# 清理验证用户
nid = user_id_by_name(admin, "verify_normal")
if nid:
    req("DELETE", "/api/admin/users/" + str(nid), token=admin)
    print("已清理 verify_normal")

print("\nALL FIELD VERIFY PASS")
