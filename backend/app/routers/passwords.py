import io
import json
import os
from datetime import datetime, timezone
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..core.deps import (
    ensure_group_access,
    get_current_user,
    get_user_group_ids,
    require_admin,
    visibility_filter,
)
from ..crypto import entry_cipher, manager
from ..db import get_db
from ..models import Group, History, OrgKey, PasswordEntry, User

router = APIRouter(
    prefix="/api/passwords",
    tags=["passwords"],
    dependencies=[Depends(get_current_user)],
)

# 条目支持的加密方式：symmetric=仅内层 SM4；gpg/sm2=外层非对称 + 内层 SM4
VALID_ALGOS = ("symmetric", "gpg", "sm2")


class CreateRequest(BaseModel):
    title: str = ""  # 密码文件名称
    username: str = ""  # 用户名 / 账号
    secret: str
    notes: str = ""
    comment: str = ""
    group_id: int  # 必填：数据绑定的分组
    system: str = ""  # 系统（如邮箱 / 服务器 / GitLab）
    algorithm: str = "symmetric"  # 'symmetric' = 条目密码对称加密；'gpg' / 'sm2' = legacy
    entry_password: str = ""  # algorithm='symmetric' 时必填；其余算法不需要
    orgkey_id: Optional[int] = None  # legacy 方案时使用：选一把本组织 OrgKey 的公钥加密


class UpdateRequest(BaseModel):
    title: Optional[str] = None
    username: Optional[str] = None
    algorithm: Optional[str] = None  # 目标算法：'symmetric' | 'gpg' | 'sm2'（省略则保持原方案）
    secret: Optional[str] = None
    notes: Optional[str] = None
    system: Optional[str] = None  # 系统
    comment: str = ""
    entry_password: Optional[str] = None  # 当前条目密码（scheme=entry 或目标改 symmetric 时必填）
    new_entry_password: Optional[str] = None  # 仅当目标为 symmetric 时使用（不填则沿用当前/服务端密钥加密）
    orgkey_id: Optional[int] = None  # legacy 方案的目标 OrgKey；省略则保持 / 回退


class UnlockRequest(BaseModel):
    """查看受「解密密码」保护的条目时，用请求体（而非 URL 查询参数）传递解密密码，
    避免密码出现在 URL、服务器访问日志与浏览器历史中。"""
    entry_password: Optional[str] = None


class ExportRequest(BaseModel):
    ids: list[int]  # 要导出的条目 id
    passwords: dict[str, str] = {}  # 明文导出时：{ "<entry_id>": "<解密密码>" }
    format: str = "xlsx"  # 仅支持 'xlsx'（Excel）
    plaintext: bool = False  # True=导出明文；False=加密备份（仅含密文，无需密码）


def _serialize_meta(db: Session, e: PasswordEntry) -> dict:
    key_name = None
    if e.orgkey_id:
        k = db.query(OrgKey).filter_by(id=e.orgkey_id).first()
        if k:
            key_name = k.name
    return {
        "id": e.id,
        "title": e.title or "",
        "system": e.system or "",
        "username": e.username,
        "algorithm": e.algorithm,
        "scheme": e.scheme,
        "needs_password": bool(e.entry_salt) and bool(e.entry_iv),
        "notes": e.notes,
        "group_id": e.group_id,
        "orgkey_id": e.orgkey_id,
        "key_name": key_name,
        "created_at": e.created_at.isoformat() if e.created_at else None,
        "updated_at": e.updated_at.isoformat() if e.updated_at else None,
        "created_by": e.created_by,
        "updated_by": e.updated_by,
    }


def _require_algorithm(algo: Optional[str]) -> Optional[str]:
    if algo is None:
        return None
    if algo not in manager.SUPPORTED:
        raise HTTPException(status_code=400, detail=f"不支持的算法: {algo}")
    return algo


def _resolve_orgkey(db: Session, user: User, orgkey_id: Optional[int], expected_group_id: int) -> Optional[OrgKey]:
    """若传了 orgkey_id：必须在用户可见组织内且与 expected_group_id 相同。"""
    if not orgkey_id:
        return None
    rec = db.query(OrgKey).filter_by(id=orgkey_id).first()
    if rec is None:
        raise HTTPException(status_code=400, detail="指定的 OrgKey 不存在")
    ensure_group_access(db, user, rec.group_id)
    if rec.group_id != expected_group_id:
        raise HTTPException(status_code=400, detail="OrgKey 与数据分组不匹配")
    return rec


def _encrypt_for_create(db: Session, user: User, req: CreateRequest) -> dict:
    """按 algorithm 分流，但无论哪种方式都先以「条目密码」做内层 SM4 加密（零知识）：

    - symmetric : 仅内层 SM4，密文直接落库（scheme='entry'）
    - gpg / sm2 : 内层 SM4 后再用 OrgKey 公钥（或服务端密钥）做一次非对称加密，
                  形成「外层非对称 + 内层对称」混合密文（scheme='legacy'）。
                  查看/修改时必须同时持有私钥（服务端）与条目密码（内层）才能还原明文。
    """
    algo = (req.algorithm or "symmetric").lower()
    if not req.entry_password:
        raise HTTPException(
            status_code=400,
            detail="无论采用哪种加密方式，新增密码都必须填写「解密密码」",
        )
    # 内层：条目密码 SM4-CBC（带 salt/iv，服务端不持久化密码本身）
    inner = entry_cipher.encrypt_entry(req.secret, req.entry_password)
    if algo == "symmetric":
        return {
            "algorithm": "symmetric",
            "scheme": "entry",
            "ciphertext": inner["ciphertext"],
            "entry_salt": inner["salt"],
            "entry_iv": inner["iv"],
            "orgkey_id": None,
        }
    if algo in ("gpg", "sm2"):
        inner_blob = json.dumps(inner, ensure_ascii=False)
        orgkey = _resolve_orgkey(db, user, req.orgkey_id, req.group_id)
        if orgkey is not None:
            try:
                ciphertext = manager.get_provider(algo).encrypt(inner_blob, orgkey.public_key)
            except Exception:
                raise HTTPException(status_code=400, detail="用 OrgKey 公钥加密失败，请检查公钥内容") from None
            return {
                "algorithm": algo,
                "scheme": "legacy",
                "ciphertext": ciphertext,
                "entry_salt": inner["salt"],
                "entry_iv": inner["iv"],
                "orgkey_id": orgkey.id,
            }
        # 回退：服务端默认密钥（兼容 OrgKey 库为空的情况）
        return {
            "algorithm": algo,
            "scheme": "legacy",
            "ciphertext": manager.encrypt_secret(db, algo, inner_blob),
            "entry_salt": inner["salt"],
            "entry_iv": inner["iv"],
            "orgkey_id": None,
        }
    raise HTTPException(status_code=400, detail=f"不支持的加密方式: {algo}")


def _legacy_decrypt(db: Session, e: PasswordEntry) -> str:
    """解开 legacy 外层（GPG/SM2），返回内层原文。

    hybrid 条目下内层是 JSON 字符串（{salt,iv,ciphertext}）；旧式纯 legacy 下内层即明文。

    支持「受口令保护的 OrgKey 私钥」：解密时会使用 OrgKey.private_passphrase 自动解锁。
    """
    if e.orgkey_id:
        try:
            return manager.decrypt_with_orgkey(db, e.orgkey_id, e.ciphertext)
        except RuntimeError:
            # 指定的 OrgKey 不存在或未持有私钥 → 回退到服务端默认密钥
            pass
        except ValueError:
            # 受口令保护但口令缺失/错误：返回脱敏文案，避免泄露内部异常细节
            raise HTTPException(
                status_code=400,
                detail="用 OrgKey 私钥解密失败（缺少或错误的解密口令）",
            ) from None
        except Exception:
            raise HTTPException(
                status_code=500,
                detail="用 OrgKey 私钥解密失败，请确认该密钥的私钥与口令是否正确",
            ) from None
    return manager.decrypt_secret(db, e.algorithm, e.ciphertext)


def _decrypt_entry_secret(db: Session, e: PasswordEntry, entry_password: Optional[str]) -> str:
    """解密单条密码明文。

    - 无内层条目密码（旧式纯 legacy）：直接用服务端/GroKey 私钥解密返回；
    - 有内层条目密码（symmetric 或 hybrid gpg/sm2）：必须提供且正确的 entry_password 才能解开内层 SM4。
    """
    has_entry = bool(e.entry_salt) and bool(e.entry_iv)
    if not has_entry:
        # 旧版 legacy：纯 GPG/SM2，无需条目密码
        return _legacy_decrypt(db, e)
    if not entry_password:
        raise HTTPException(
            status_code=400,
            detail="该密码由「解密密码」保护，请提供 entry_password 才能查看",
        )
    try:
        if e.scheme == "entry":
            plaintext = entry_cipher.decrypt_entry(
                {"salt": e.entry_salt, "iv": e.entry_iv, "ciphertext": e.ciphertext},
                entry_password,
            )
        else:
            inner_blob = _legacy_decrypt(db, e)  # hybrid：外层 GPG/SM2 -> 内层 JSON
            inner = json.loads(inner_blob)
            plaintext = entry_cipher.decrypt_entry(inner, entry_password)
    except entry_cipher.WrongPasswordError:
        raise HTTPException(status_code=401, detail="解密密码错误，无法解密")
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="密文格式损坏，无法解析内层数据")
    return plaintext


@router.get("")
def list_passwords(
    page: int = None,
    page_size: int = None,
    q: Optional[str] = None,
    group_id: Optional[int] = None,
    system: Optional[str] = None,
    db: Session = Depends(get_db), user: User = Depends(get_current_user)
):
    """密码列表。

    - 不传 page_size：返回完整扁平数组（兼容旧调用 / 其它内部页面）。
    - 传 page_size：返回分页信封 {"items", "total", "page", "page_size"}（密码页走此路径）。
    - q：按「账号 / 密码文件名称 / 系统 / 备注」模糊搜索（后台执行）。
    - group_id / system：可选筛选（分组下拉 + 系统输入框）。
    """
    gids = get_user_group_ids(db, user)
    f = visibility_filter(PasswordEntry.group_id, user, gids)
    qry = db.query(PasswordEntry).filter_by(deleted=False)
    if group_id is not None:
        qry = qry.filter(PasswordEntry.group_id == group_id)
    if system:
        qry = qry.filter(PasswordEntry.system.ilike(f"%{system.strip()}%"))
    if f is not None:
        qry = qry.filter(f)
    rows = qry.order_by(PasswordEntry.updated_at.desc()).all()
    if q:
        ql = q.strip().lower()
        rows = [
            r for r in rows
            if ql in (r.username or "").lower()
            or ql in (r.title or "").lower()
            or ql in (r.notes or "").lower()
            or ql in (r.system or "").lower()
        ]
    total = len(rows)
    # 未指定分页 → 兼容旧调用：返回完整扁平数组
    if page_size is None:
        return [_serialize_meta(db, r) for r in rows]
    # 指定分页 → 返回信封
    page = page or 1
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 20
    if page_size > 5000:
        page_size = 5000
    total_pages = (total + page_size - 1) // page_size or 1
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    page_rows = rows[start:start + page_size]
    items = [_serialize_meta(db, r) for r in page_rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.post("")
def create(
    req: CreateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    ensure_group_access(db, user, req.group_id)

    # 重复新增校验：同一分组下不允许「账号名称 + 加密方式」完全相同
    existing = (
        db.query(PasswordEntry)
        .filter_by(group_id=req.group_id, algorithm=req.algorithm, deleted=False)
        .all()
    )
    u_name = (req.username or "").strip().lower()
    for r in existing:
        if (r.username or "").strip().lower() == u_name:
            raise HTTPException(
                status_code=409,
                detail=f"该分组下已存在账号「{req.username}」且加密方式相同（{req.algorithm}），请勿重复新增",
            )

    fields = _encrypt_for_create(db, user, req)
    entry = PasswordEntry(
        title=(req.title or "").strip(),
        username=req.username,
        notes=req.notes,
        system=(req.system or "").strip(),
        group_id=req.group_id,
        created_by=user.username,
        updated_by=user.username,
        **fields,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    db.add(
        History(
            password_id=entry.id,
            group_id=entry.group_id,
            action="create",
            title=entry.title,
            username=entry.username,
            system=entry.system,
            algorithm=entry.algorithm,
            ciphertext=entry.ciphertext,
            notes=entry.notes,
            changed_by=user.username,
            comment=req.comment or "新增密码",
        )
    )
    db.commit()
    return {"id": entry.id, "message": "created"}



# ────────────────────── 批量导入密码：模板下载 + 上传解析 ──────────────────────
# 设计：加密方式 / 加密密码（解密密码）/ 密钥 在「导入页面」上统一选择，对所有行生效；
# 每一行只需提供 用户名 / 密码明文（真实密码）/ 备注 / 所属分组。
# 与用户批量导入一致：逐行解析、逐行创建，某行失败不影响其它行，并以逐行报告回执。

PWD_IMPORT_HEADERS = ["密码文件名称", "系统", "用户名", "密码明文", "备注"]


def _xlsx_bytes_passwords() -> bytes:
    """用 openpyxl 渲染密码导入模板到 bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "密码批量导入模板"

    bold = Font(bold=True)
    head_fill = PatternFill(start_color="FFEAF2FF", end_color="FFEAF2FF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="使用说明").font = Font(bold=True, color="FF2563EB")
    instructions = [
        "本表用于批量导入密码。第 1 行为说明，导入时会自动忽略。",
        "表头为：密码文件名称 / 系统 / 用户名 / 密码明文 / 备注，其下方即为数据行。",
        "「密码文件名称」为该条密码的命名（如「邮箱登录密码」）；「系统」为所属系统（如「邮箱系统 / 生产服务器」，可留空）；「用户名」为登录账号 / 邮箱；「密码明文」即你要保存的真实密码 / 密钥本身；「备注」为可选说明。",
        "「密码明文」与导入页面填写的「加密密码（解密密码）」是两回事 —— 加密密码仅用于解锁本批导入的条目。",
        "「所属分组」请在导入页面的下拉框中选择（对所有行生效），无需在模板里填写。",
        "加密方式 / 加密密码 / 密钥在导入页面上统一选择，对所有行生效。",
        "下方两行示例数据仅作演示，上传前请整行删除。",
    ]
    for i, line in enumerate(instructions, start=2):
        ws.cell(row=i, column=1, value=line)

    note_end = 1 + len(instructions)
    body_start = note_end + 2  # 与说明之间留一行空行，再放表头

    for col_idx, header in enumerate(PWD_IMPORT_HEADERS, start=1):
        c = ws.cell(row=body_start, column=col_idx, value=header)
        c.font = bold
        c.fill = head_fill
        c.alignment = center

    examples = [
        ["邮箱登录密码", "邮箱系统", "alice@example.com", "Alice@2026", "示例"],
        ["服务器root", "生产服务器", "root", "Root@2026!", "示例"],
    ]
    for r, row in enumerate(examples, start=body_start + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r, column=c_idx, value=val).alignment = Alignment(vertical="center")

    for col_idx, header in enumerate(PWD_IMPORT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) * 2 + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/template")
def download_password_template(
    fmt: str = "xlsx",
    _: User = Depends(get_current_user),
):
    """下载批量导入密码的模板，仅支持 Excel (.xlsx) 格式。"""
    fmt = (fmt or "xlsx").lower()
    if fmt != "xlsx":
        raise HTTPException(status_code=400, detail="导入模板仅支持 Excel (.xlsx) 格式")
    data = _xlsx_bytes_passwords()
    filename = "密码批量导入模板.xlsx"
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    quoted = quote(filename)
    headers = {
        "Content-Disposition": (
            f"attachment; filename=\"password_import_template.xlsx\"; filename*=UTF-8''{quoted}"
        )
    }
    return StreamingResponse(io.BytesIO(data), media_type=media, headers=headers)


# ────────────────────── 上传解析 ──────────────────────

def _norm_pwd_row(row: dict) -> tuple:
    """把一行数据归一化成 (title, system, username, secret, notes)。所属分组由页面统一选择，不在模板内。"""
    title = (row.get("title") or "").strip()
    system = (row.get("system") or "").strip()
    username = (row.get("username") or "").strip()
    secret = (row.get("secret") or "").strip()  # 密码明文（真实密码）
    notes = (row.get("notes") or "").strip()
    return title, system, username, secret, notes


# 表头别名：兼容新模板（密码文件名称/系统/用户名）与旧导出文件（标题/账号）
PWD_HEADER_ALIASES = {
    "title": ["密码文件名称", "标题"],
    "system": ["系统"],
    "username": ["用户名", "账号"],
    "secret": ["密码明文"],
    "notes": ["备注"],
}


def _read_xlsx_passwords(content: bytes) -> List[dict]:
    """从 xlsx bytes 提取数据行；返回 dict 列表（key 是逻辑字段名）。

    支持新表头（密码文件名称 / 系统 / 用户名 / 密码明文 / 备注）与旧表头别名
    （标题 / 账号），以便历史导出的文件也能重新导入。
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers = None
    out: List[dict] = []
    for r in ws.iter_rows(values_only=True):
        row = [c for c in r if c is not None]
        if not row:
            continue
        first = str(row[0]).strip()
        # 跳过说明行
        if first.startswith("#") or first.startswith("使用说明") or first.startswith("下方为"):
            continue
        if headers is None:
            raw_headers = [str(x).strip() for x in row]
            if "密码明文" not in raw_headers:
                continue
            headers = {}
            for key, aliases in PWD_HEADER_ALIASES.items():
                for alias in aliases:
                    if alias in raw_headers:
                        headers[key] = alias
                        break
            continue
        full = list(r)
        cells = {}
        for key, header_name in headers.items():
            try:
                col_idx = raw_headers.index(header_name)
            except ValueError:
                continue
            cells[key] = "" if col_idx >= len(full) or full[col_idx] is None else str(full[col_idx])
        out.append(cells)
    wb.close()
    return out


class PasswordImportRow(BaseModel):
    row: int
    username: str
    status: str  # "created" | "skipped" | "error"
    message: str = ""


class PasswordImportResponse(BaseModel):
    total: int
    created: int
    skipped: int
    errored: int
    rows: List[PasswordImportRow]


@router.post("/import", response_model=PasswordImportResponse)
async def import_passwords(
    file: UploadFile = File(...),
    algorithm: str = Form("symmetric"),
    entry_password: str = Form(""),
    orgkey_id: Optional[int] = Form(None),
    group_id: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量导入密码：上传 .xlsx，按页面选择的加密方式 / 加密密码 / 密钥逐行创建。

    - ``algorithm``    ：加密方式（symmetric / gpg / sm2），对所有行生效
    - ``entry_password``：加密密码（即「解密密码」），对所有行生效，必填
    - ``orgkey_id``    ：gpg / sm2 时可选的 OrgKey（与每行分组须匹配，否则该行报错）
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="文件为空")

    max_bytes = int(os.getenv("BATCH_UPLOAD_MAX_BYTES", "10485760"))  # 默认 10MB
    if len(raw) > max_bytes:
        raise HTTPException(status_code=413, detail="上传文件过大（上限 10MB）")

    fname = (file.filename or "").lower()
    if not fname.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx（Excel）文件")

    try:
        data_rows = _read_xlsx_passwords(raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析 xlsx：{e}")

    if not data_rows:
        raise HTTPException(
            status_code=400,
            detail="未找到可导入的数据行；请检查表头是否为「账号 / 密码明文 / 备注」",
        )

    if not entry_password:
        raise HTTPException(status_code=400, detail="请先在页面选择「加密密码（解密密码）」再导入")

    # 校验页面选择的所属分组（对所有导入行生效）
    g = db.query(Group).filter_by(id=group_id).first()
    if not g:
        raise HTTPException(status_code=400, detail=f"所属分组（id={group_id}）不存在")
    try:
        ensure_group_access(db, user, group_id)
    except HTTPException:
        raise HTTPException(status_code=403, detail="无权访问所选分组")

    algo = (algorithm or "symmetric").lower()
    if algo not in VALID_ALGOS:
        raise HTTPException(status_code=400, detail=f"不支持的加密方式: {algo}")

    results: List[PasswordImportRow] = []
    created = skipped = errored = 0

    for idx, raw_row in enumerate(data_rows, start=1):
        title, system, username, secret, notes = _norm_pwd_row(raw_row)
        if not username and not secret:
            results.append(PasswordImportRow(row=idx, username="", status="skipped", message="空行已跳过"))
            skipped += 1
            continue

        if not username:
            results.append(PasswordImportRow(row=idx, username="", status="error", message="账号为空"))
            errored += 1
            continue
        if not secret:
            results.append(PasswordImportRow(row=idx, username=username, status="error", message="密码明文为空"))
            errored += 1
            continue

        # 重复校验：同一分组下「账号 + 加密方式」相同则跳过
        if db.query(PasswordEntry).filter_by(
            group_id=group_id, algorithm=algo, username=username, deleted=False
        ).first():
            results.append(PasswordImportRow(
                row=idx, username=username, status="error",
                message="该分组下已存在相同账号与加密方式",
            ))
            errored += 1
            continue

        req = CreateRequest(
            title=title or username,
            system=system,
            username=username,
            secret=secret,
            notes=notes,
            group_id=group_id,
            algorithm=algo,
            entry_password=entry_password,
            orgkey_id=orgkey_id if algo != "symmetric" else None,
        )
        try:
            fields = _encrypt_for_create(db, user, req)
        except HTTPException as e:
            results.append(PasswordImportRow(row=idx, username=username, status="error", message=e.detail))
            errored += 1
            continue

        entry = PasswordEntry(
            title=title or username,
            system=system,
            username=username,
            notes=notes,
            group_id=group_id,
            created_by=user.username,
            updated_by=user.username,
            **fields,
        )
        db.add(entry)
        try:
            db.flush()
        except Exception as e:
            db.rollback()
            results.append(PasswordImportRow(row=idx, username=username, status="error", message=f"数据库写入失败：{e}"))
            errored += 1
            continue

        db.add(History(
            password_id=entry.id,
            group_id=group_id,
            action="create",
            title=username,
            username=username,
            algorithm=algo,
            ciphertext=entry.ciphertext,
            notes=notes,
            changed_by=user.username,
            comment="批量导入",
        ))
        db.commit()
        created += 1
        msg = "已创建"
        results.append(PasswordImportRow(row=idx, username=username, status="created", message=msg))

    return PasswordImportResponse(
        total=len(data_rows),
        created=created,
        skipped=skipped,
        errored=errored,
        rows=results,
    )

@router.get("/{pid}")
def get_one(
    pid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """无「解密密码」的旧式条目（纯 GPG/SM2 服务端密钥加密）可直接查看；
    受「解密密码」保护的条目请使用 POST /{pid}/unlock 并在请求体中传密码。"""
    entry = db.query(PasswordEntry).filter_by(id=pid, deleted=False).first()
    if entry is None:
        raise HTTPException(status_code=404, detail="未找到该密码")
    ensure_group_access(db, user, entry.group_id)
    if bool(entry.entry_salt) and bool(entry.entry_iv):
        raise HTTPException(
            status_code=400,
            detail="该密码由「解密密码」保护，请使用 POST /api/passwords/{id}/unlock 并在请求体中传 entry_password",
        )
    secret = _decrypt_entry_secret(db, entry, None)
    return {**_serialize_meta(db, entry), "secret": secret}


@router.post("/{pid}/unlock")
def unlock(
    pid: int,
    req: UnlockRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """受「解密密码」保护的条目：用请求体（JSON）传 entry_password，解密后返回明文。
    密码只在请求体内传输，不会出现在 URL / 访问日志 / 浏览器历史中。"""
    entry = db.query(PasswordEntry).filter_by(id=pid, deleted=False).first()
    if entry is None:
        raise HTTPException(status_code=404, detail="未找到该密码")
    ensure_group_access(db, user, entry.group_id)
    secret = _decrypt_entry_secret(db, entry, req.entry_password)
    return {**_serialize_meta(db, entry), "secret": secret}


@router.put("/{pid}")
def update(
    pid: int,
    req: UpdateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    entry = db.query(PasswordEntry).filter_by(id=pid, deleted=False).first()
    if entry is None:
        raise HTTPException(status_code=404, detail="未找到该密码")
    ensure_group_access(db, user, entry.group_id)

    changes: list[str] = []
    if req.title is not None and req.title != entry.title:
        entry.title = req.title
        changes.append("title")
    if req.username is not None and req.username != entry.username:
        entry.username = req.username
        changes.append("username")
    if req.notes is not None and req.notes != entry.notes:
        entry.notes = req.notes
        changes.append("备注")
    if req.system is not None and req.system != entry.system:
        entry.system = req.system
        changes.append("system")
    if req.orgkey_id is not None and req.orgkey_id != entry.orgkey_id:
        changes.append("加密密钥")

    has_entry = bool(entry.entry_salt) and bool(entry.entry_iv)
    target_algo = (req.algorithm or entry.algorithm or "symmetric").lower()
    target_scheme = "entry" if target_algo == "symmetric" else "legacy"
    algo_changed = target_algo != entry.algorithm or (
        ("entry" if entry.algorithm == "symmetric" else "legacy") != target_scheme
    )

    # 旧式纯 legacy（无解密密码层）且目标仍为 legacy 且未提供新解密密码 -> 维持旧式（不引入条目密码层）
    preserve_noentry = (not has_entry) and target_algo in ("gpg", "sm2") and not (
        req.new_entry_password or req.entry_password
    )

    # ---- 读取当前明文 ----
    if preserve_noentry:
        current_secret = _legacy_decrypt(db, entry)
    else:
        if has_entry:
            if not req.entry_password:
                raise HTTPException(
                    status_code=400,
                    detail="修改受「解密密码」保护的密码必须提供 entry_password",
                )
            try:
                if entry.scheme == "entry":
                    current_secret = entry_cipher.decrypt_entry(
                        {"salt": entry.entry_salt, "iv": entry.entry_iv, "ciphertext": entry.ciphertext},
                        req.entry_password,
                    )
                else:
                    inner_blob = _legacy_decrypt(db, entry)
                    current_secret = entry_cipher.decrypt_entry(json.loads(inner_blob), req.entry_password)
            except entry_cipher.WrongPasswordError:
                raise HTTPException(status_code=401, detail="解密密码错误，无法修改")
        else:
            # 旧式 legacy 升级：无需密码即可读取，但写入时必须指定解密密码
            current_secret = _legacy_decrypt(db, entry)

    # 仅在明确提供了非空新明文时才替换；空串视为「保持不变」，
    # 避免编辑标题/备注时因前端不预填明文而误将密码清空。
    if req.secret not in (None, ""):
        new_secret = req.secret
        if req.secret != current_secret:
            changes.append("密码明文")
    else:
        new_secret = current_secret

    if target_algo not in VALID_ALGOS:
        raise HTTPException(status_code=400, detail=f"不支持的加密方式: {target_algo}")

    # ---- 维持旧式（纯 GPG/SM2，无解密密码层）----
    if preserve_noentry:
        if new_secret != current_secret or (req.orgkey_id is not None and req.orgkey_id != entry.orgkey_id):
            orgkey = _resolve_orgkey(db, user, req.orgkey_id, entry.group_id) if req.orgkey_id is not None else None
            if orgkey is not None:
                entry.ciphertext = manager.get_provider(target_algo).encrypt(new_secret, orgkey.public_key)
                entry.orgkey_id = orgkey.id
            else:
                entry.ciphertext = manager.encrypt_secret(db, target_algo, new_secret)
                entry.orgkey_id = None
            entry.entry_salt = ""
            entry.entry_iv = ""
        entry.algorithm = target_algo
        entry.scheme = "legacy"
    else:
        # ---- 新设计：始终带「解密密码」内层 ----
        enc_pw = req.new_entry_password or req.entry_password
        if not enc_pw:
            raise HTTPException(
                status_code=400,
                detail="必须提供解密密码（或新解密密码）才能保存",
            )
        inner = entry_cipher.encrypt_entry(new_secret, enc_pw)
        if target_scheme == "entry":
            entry.algorithm = "symmetric"
            entry.scheme = "entry"
            entry.ciphertext = inner["ciphertext"]
            entry.entry_salt = inner["salt"]
            entry.entry_iv = inner["iv"]
            entry.orgkey_id = None
        else:
            inner_blob = json.dumps(inner, ensure_ascii=False)
            orgkey = _resolve_orgkey(db, user, req.orgkey_id, entry.group_id)
            if orgkey is not None:
                try:
                    entry.ciphertext = manager.get_provider(target_algo).encrypt(inner_blob, orgkey.public_key)
                except Exception as e:
                    raise HTTPException(status_code=400, detail=f"用 OrgKey 公钥加密失败：{e}") from e
                entry.orgkey_id = orgkey.id
            else:
                entry.ciphertext = manager.encrypt_secret(db, target_algo, inner_blob)
                entry.orgkey_id = None
            entry.algorithm = target_algo
            entry.scheme = "legacy"
            entry.entry_salt = inner["salt"]
            entry.entry_iv = inner["iv"]

    if (not preserve_noentry) and req.new_entry_password:
        changes.append("解密密码")
    if algo_changed:
        changes.append("加密方式")
    entry.updated_by = user.username
    entry.updated_at = datetime.now(timezone.utc)
    db.commit()

    db.add(
        History(
            password_id=entry.id,
            group_id=entry.group_id,
            action="update",
            title=entry.title,
            username=entry.username,
            system=entry.system,
            algorithm=entry.algorithm,
            ciphertext=entry.ciphertext,
            notes=entry.notes,
            changed_by=user.username,
            comment=req.comment or ("修改了 " + "，".join(changes) if changes else "无变更"),
        )
    )
    db.commit()
    return {"id": pid, "message": "updated", "changes": changes}


@router.delete("/{pid}")
def delete(
    pid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    entry = db.query(PasswordEntry).filter_by(id=pid, deleted=False).first()
    if entry is None:
        raise HTTPException(status_code=404, detail="未找到该密码")
    ensure_group_access(db, user, entry.group_id)
    entry.deleted = True
    entry.updated_by = user.username
    entry.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.add(
        History(
            password_id=pid,
            group_id=entry.group_id,
            action="delete",
            title=entry.title,
            username=entry.username,
            system=entry.system,
            algorithm=entry.algorithm,
            ciphertext=entry.ciphertext,
            notes=entry.notes,
            changed_by=user.username,
            comment=f"删除密码（账号：{entry.username or entry.title or '未命名'}）",
        )
    )
    db.commit()
    return {"id": pid, "message": "deleted"}


def _build_export_rows(db: Session, entries: list, plaintext: bool, passwords: dict) -> tuple:
    """构造导出数据行。返回 (rows, skipped, failed)。

    - plaintext=False（加密备份）：仅含密文与元数据，无需密码，可用于迁移 / 恢复。
    - plaintext=True：尝试用 passwords 中的解密密码还原明文；失败则该条 secret=None、
      计入 skipped，并记录其账号名到 failed，供上层在明文导出时整体拒绝并提示。
    """
    groups = {g.id: g.name for g in db.query(Group).all()}
    rows = []
    skipped = 0
    failed = []
    for e in entries:
        key_name = None
        if e.orgkey_id:
            k = db.query(OrgKey).filter_by(id=e.orgkey_id).first()
            if k:
                key_name = k.name
        has_entry = bool(e.entry_salt) and bool(e.entry_iv)
        row = {
            "id": e.id,
            "title": e.title or "",
            "system": e.system or "",
            "username": e.username,
            "algorithm": e.algorithm,
            "scheme": e.scheme,
            "group_id": e.group_id,
            "group_name": groups.get(e.group_id, "—"),
            "key_name": key_name,
            "notes": e.notes or "",
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "updated_at": e.updated_at.isoformat() if e.updated_at else None,
        }
        if plaintext:
            pw = passwords.get(str(e.id)) or passwords.get(e.id)
            secret = None
            decrypt_error = False
            if not has_entry:
                # 旧式纯 legacy：服务端密钥可直接还原
                try:
                    secret = _legacy_decrypt(db, e)
                except Exception:
                    decrypt_error = True
                    skipped += 1
                    failed.append(e.username or f"#{e.id}")
            else:
                try:
                    secret = _decrypt_entry_secret(db, e, pw)
                except Exception:
                    decrypt_error = True
                    skipped += 1
                    failed.append(e.username or f"#{e.id}")
            row["secret"] = secret
            row["decrypt_error"] = decrypt_error
        else:
            # 加密备份：仅导出密文，不含明文
            row["ciphertext"] = e.ciphertext
            row["entry_salt"] = e.entry_salt or ""
            row["entry_iv"] = e.entry_iv or ""
            row["orgkey_id"] = e.orgkey_id
        rows.append(row)
    return rows, skipped, failed


def _export_filename(fmt: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"密码导出_{ts}.xlsx"


def _xlsx_bytes_export(rows: list, plaintext: bool, skipped: int) -> bytes:
    """把导出数据行渲染成 .xlsx 字节。

    - plaintext=True ：输出「账号 / 加密方式 / 所属分组 / 密钥 / 密码明文 / 备注 / 更新时间」
    - plaintext=False：输出「账号 / 加密方式 / 所属分组 / 密钥 / 密文 / 备注 / 更新时间」（仅含密文，用于迁移/恢复）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "密码导出"

    bold = Font(bold=True)
    head_fill = PatternFill(start_color="FFEAF2FF", end_color="FFEAF2FF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    if plaintext:
        headers = ["密码文件名称", "系统", "用户名", "加密方式", "所属分组", "密钥", "密码明文", "备注", "更新时间"]
    else:
        headers = ["密码文件名称", "系统", "用户名", "加密方式", "所属分组", "密钥", "密文", "备注", "更新时间"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = bold
        c.fill = head_fill
        c.alignment = center

    for r, row in enumerate(rows, start=2):
        if plaintext:
            if row.get("decrypt_error"):
                secret = "解密失败"
            elif row.get("secret") is None:
                secret = ""
            else:
                secret = row.get("secret")
            vals = [
                row.get("title", ""),
                row.get("system", ""),
                row.get("username", ""),
                row.get("algorithm", ""),
                row.get("group_name", ""),
                row.get("key_name") or "",
                secret,
                row.get("notes", ""),
                row.get("updated_at") or "",
            ]
        else:
            vals = [
                row.get("title", ""),
                row.get("system", ""),
                row.get("username", ""),
                row.get("algorithm", ""),
                row.get("group_name", ""),
                row.get("key_name") or "",
                row.get("ciphertext", ""),
                row.get("notes", ""),
                row.get("updated_at") or "",
            ]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="center")

    # 列宽自适应
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        name = get_column_letter(col_idx)
        maxlen = len(str(headers[col_idx - 1]))
        for r in range(2, len(rows) + 2):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[name].width = min(60, max(14, maxlen + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/export")
def export_passwords(
    req: ExportRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量导出所选密码（所有登录用户可用），仅支持 Excel (.xlsx) 格式。

    - plaintext=False：加密备份（含密文 + 元数据），无需密码，可用于迁移 / 恢复。
    - plaintext=True：明文导出，需通过 passwords 提供各条目的解密密码；解密失败者 secret 为 null。
    解密密码通过请求体（JSON）传输，不会出现在 URL 中。
    """
    fmt = (req.format or "xlsx").lower()
    if fmt != "xlsx":
        raise HTTPException(status_code=400, detail="导出仅支持 Excel (.xlsx) 格式")
    if not req.ids:
        raise HTTPException(status_code=400, detail="请至少选择一条密码")

    gids = get_user_group_ids(db, user)
    f = visibility_filter(PasswordEntry.group_id, user, gids)
    q = db.query(PasswordEntry).filter_by(deleted=False).filter(PasswordEntry.id.in_(req.ids))
    if f is not None:
        q = q.filter(f)
    entries = q.all()
    if not entries:
        raise HTTPException(status_code=404, detail="没有可导出的可见条目")

    rows, skipped, failed = _build_export_rows(db, entries, req.plaintext, req.passwords or {})

    # 明文导出：只要有任一条目解密失败，整体拒绝导出并给出提示，避免导出残缺/空密文
    if req.plaintext and skipped > 0:
        names = "、".join(failed) if failed else f"{skipped} 条"
        raise HTTPException(
            status_code=400,
            detail=f"有 {skipped} 条密码解密失败，无法导出明文（{names}）。请检查这些条目的解密密码是否正确。",
        )

    # 安全审计：明文导出属于高风险操作（相当于整批导出凭据），记入审计日志以便追溯
    if req.plaintext:
        db.add(
            History(
                password_id=None,
                group_id=entries[0].group_id if entries else None,
                action="export",
                title="明文导出密码",
                username=None,
                algorithm=None,
                ciphertext=None,
                notes=f"明文导出 {len(entries)} 条密码",
                changed_by=user.username,
                comment=f"明文导出 {len(entries)} 条密码",
            )
        )
        db.commit()

    content = _xlsx_bytes_export(rows, req.plaintext, skipped)
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    fname = _export_filename(fmt)
    ascii_name = f"password_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    disposition = f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(fname)}'
    return Response(content=content, media_type=media_type, headers={"Content-Disposition": disposition})

