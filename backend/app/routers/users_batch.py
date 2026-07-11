"""批量新增用户：Excel 模板下载 + 解析上传 + 部分失败回执。

设计目标
=========
* 管理员可先下载 Excel 模板（.xlsx），照表头填写后上传；
* 仅支持 ``.xlsx`` 格式（用 openpyxl 生成与解析），不再支持 CSV。
* 上传后逐行解析、逐行创建；某行失败不影响其它行。响应里逐行列出「成功 / 失败原因」。
* 单密码字段强度不做强制（管理员场景），但必须非空。
"""
from __future__ import annotations

import io
import os
from typing import List, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..core.deps import require_admin, ensure_group_access
from ..db import get_db
from ..models import Group, User, user_groups
from ..security import derive_password_verifier, hash_password
from .admin import _link_user_group, _seed_login_material

router = APIRouter(prefix="/api/admin/users", tags=["admin-users-batch"])


# ────────────────────── 模板（仅 xlsx） ──────────────────────

# Excel 模板表头（中英结合方便阅读，但解析时按中文字段识别）
TEMPLATE_HEADERS = ["用户名", "密码", "是否管理员"]


def _build_template_rows() -> List[List[str]]:
    """生成示例行：前 2 行为示例 + 表头说明，最后 1 行为管理员示例。"""
    return [
        # 下面是示例，可保留也可以全部删除后整列上传
        ["alice", "Alice@2026", "否"],
        ["bob", "BobSecret!9", "否"],
        ["carol", "Carol#Admin1", "是"],
    ]


def _xlsx_bytes(rows: List[List[str]], sheet_name: str = "用户批量导入模板") -> bytes:
    """用 openpyxl 渲染 xlsx 到 bytes（无需保存到磁盘）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    bold = Font(bold=True)
    head_fill = PatternFill(start_color="FFEAF2FF", end_color="FFEAF2FF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # 第 1 行放使用说明
    ws.cell(row=1, column=1, value="使用说明").font = Font(bold=True, color="FF2563EB")
    instructions = [
        "1) 第 1 行为使用说明，解析时会忽略；请勿在下方插入空行。",
        "2) 第 2 行为表头；所属分组请在上传页面的下拉框中选择，无需在模板里填写。",
        "3) 第 3 行起是示例数据，仅作演示；管理员上传前请先清空示例行或保留作为格式参考。",
        "4) 「是否管理员」请填：是 / 否（不区分大小写）。其它值按「否」处理。",
        "5) 用户名重复、密码为空都会导致该行失败，其它行不受影响；",
        "   上传完成后会返回包含逐行结果的报告。",
    ]
    for i, line in enumerate(instructions, start=2):
        ws.cell(row=i, column=1, value=line)

    # 留一行空行
    note_end = 1 + len(instructions)
    ws.cell(row=note_end + 2, column=1, value="下方为模板数据（请在此区域内填写）：").font = Font(bold=True)

    body_start = note_end + 4  # 标题行

    # 表头
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=body_start, column=col_idx, value=header)
        c.font = bold
        c.fill = head_fill
        c.alignment = center

    # 主体行（示例）
    for r, row in enumerate(_build_template_rows(), start=body_start + 1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="center")

    # 列宽
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        width = max(14, len(str(header)) * 2 + 4)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 保存到内存
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/template")
def download_template(
    _: User = Depends(require_admin),
):
    """下载批量导入用户的 Excel 模板（.xlsx）。"""
    data = _xlsx_bytes([])
    filename = "用户批量导入模板.xlsx"
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 双字段 Content-Disposition：ASCII 兼容 + UTF-8 兼容
    from urllib.parse import quote
    quoted = quote(filename)
    headers = {
        "Content-Disposition": (
            f"attachment; filename=\"import_template.xlsx\"; filename*=UTF-8''{quoted}"
        )
    }
    return StreamingResponse(io.BytesIO(data), media_type=media, headers=headers)


# ────────────────────── 上传解析 ──────────────────────

def _norm_row(row: dict) -> Tuple[str, str, bool]:
    """把一行数据归一化成 (username, password, is_admin)。所属分组由页面统一选择，不在模板内。"""
    username = (row.get("username") or "").strip()
    password = (row.get("password") or "").strip()
    admin_raw = (row.get("is_admin") or "").strip().lower()
    is_admin = admin_raw in ("是", "yes", "y", "true", "1", "管理员")
    return username, password, is_admin


def _read_xlsx(content: bytes) -> List[dict]:
    """从 xlsx bytes 提取行；返回 dict 列表（key 是中文字段）。"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = None
    out: List[dict] = []
    for r in rows_iter:
        row = [c for c in r if c is not None]
        if not row:
            continue
        first = str(row[0]).strip() if row else ""
        # 跳过说明行（# 开头 / 「使用说明」/「下方为模板数据…」）
        if first.startswith("#") or first.startswith("使用说明") or first.startswith("下方为"):
            continue
        if headers is None:
            # 把表头键统一为归一化字段标识
            raw_headers = [str(x).strip() for x in row]
            if "用户名" not in raw_headers:
                # 表头未识别，跳过该行继续找
                continue
            headers = {
                "username": "用户名",
                "password": "密码",
                "is_admin": "是否管理员",
            }
            continue
        # 普通数据行：用 raw_headers index 对齐
        # 重新取一次完整行（包括 None），以便按列名索引
        full = list(r)
        # 同一列号取单元格
        cells = {}
        for key, header_name in headers.items():
            try:
                col_idx = raw_headers.index(header_name)
            except ValueError:
                continue
            if col_idx < len(full):
                cells[key] = "" if full[col_idx] is None else str(full[col_idx])
            else:
                cells[key] = ""
        out.append(cells)
    wb.close()
    return out


class BatchResultRow(BaseModel):
    row: int
    username: str
    status: str  # "created" | "skipped" | "error"
    message: str = ""


class BatchResponse(BaseModel):
    total: int
    created: int
    skipped: int
    errored: int
    rows: List[BatchResultRow]


@router.post("/batch", response_model=BatchResponse)
async def batch_import_users(
    file: UploadFile = File(...),
    group_ids: List[int] = Form([]),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """批量导入用户：上传 .xlsx 文件，逐行创建用户并报告每行的结果。"""
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="文件为空")

    # 上传体积上限（防 zip-bomb / 内存耗尽型 DoS）
    max_bytes = int(os.getenv("BATCH_UPLOAD_MAX_BYTES", "10485760"))  # 默认 10MB
    if len(raw) > max_bytes:
        raise HTTPException(status_code=413, detail="上传文件过大（上限 10MB）")

    fname = (file.filename or "").lower()
    if fname.endswith(".xlsx"):
        try:
            data_rows = _read_xlsx(raw)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"无法解析 xlsx：{e}")
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    if not data_rows:
        raise HTTPException(status_code=400, detail="未找到可导入的数据行；请检查表头是否为「用户名 / 密码 / 是否管理员」")

    # 校验页面选择的所属分组（为空表示不绑定任何分组）
    for gid in group_ids:
        g = db.query(Group).filter_by(id=gid).first()
        if not g:
            raise HTTPException(status_code=400, detail=f"所属分组（id={gid}）不存在")
        try:
            ensure_group_access(db, user, gid)
        except HTTPException:
            raise HTTPException(status_code=403, detail="无权访问所选分组")

    results: List[BatchResultRow] = []
    created = skipped = errored = 0

    for idx, raw_row in enumerate(data_rows, start=1):
        username, password, is_admin = _norm_row(raw_row)
        if not username and not password:
            # 空行：跳过，计入 skipped（不报错，避免模板示例行报错）
            results.append(BatchResultRow(row=idx, username="", status="skipped", message="空行已跳过"))
            skipped += 1
            continue

        if not username:
            results.append(BatchResultRow(row=idx, username="", status="error", message="用户名为空"))
            errored += 1
            continue
        if not password:
            results.append(BatchResultRow(row=idx, username=username, status="error", message="密码为空"))
            errored += 1
            continue

        if db.query(User).filter_by(username=username).first():
            results.append(BatchResultRow(row=idx, username=username, status="error", message="用户名已存在"))
            errored += 1
            continue

        # 所属分组来自页面选择（group_ids 参数）；为空表示不绑定任何分组
        warn = ""

        # 创建用户 + SCRAM 凭据
        new_user = User(
            username=username,
            hashed_password=hash_password(password),
            is_admin=is_admin,
        )
        _seed_login_material(new_user, password)
        db.add(new_user)
        try:
            db.flush()
        except Exception as e:
            db.rollback()
            results.append(BatchResultRow(row=idx, username=username, status="error", message=f"数据库写入失败：{e}"))
            errored += 1
            continue

        for gid in group_ids:
            _link_user_group(db, new_user.id, gid)
        db.commit()
        created += 1
        msg = "已创建" + (f"（{warn}）" if warn else "")
        results.append(BatchResultRow(row=idx, username=username, status="created", message=msg))

    return BatchResponse(
        total=len(data_rows),
        created=created,
        skipped=skipped,
        errored=errored,
        rows=results,
    )
